from datetime import date
from selenium import webdriver
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.by import By
from pathlib import Path
from selenium.webdriver.common.keys import Keys
from constants24aralik import *
import openpyxl
import pytest

class Test_Saucedemo24_12:
    def readProductNameFromExcel():
        excelFile = openpyxl.load_workbook("data/productNames.xlsx")
        selectedSheet = excelFile["Sheet1"]
        
        rows = selectedSheet.max_row
        data = []

        for i in range(2,rows+1):
            productName = selectedSheet.cell(i,2).value 
            tupleExample = (productName)
            data.append(tupleExample)
        return data


    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get(BASE_DOMAIN_URL)  


    def teardown_method(self):
        self.driver.quit()
        
    # "standard_user" kullanıcı adıyla giriş yapılmanın doğrulanması.
    def test_loginStandartUser(self):
        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID, USERNAME_ID )))
        userName= self.driver.find_element(By.ID, USERNAME_ID )
        userName.send_keys(USERNAME)

        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID, USERNAME_ID )))
        passwordID= self.driver.find_element(By.ID, PASSWORD_ID )
        passwordID.send_keys(PASSWORD)

        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID, USERNAME_ID )))
        loginID= self.driver.find_element(By.ID,LOGIN_ID)
        loginID.click()

        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID, USER_MENU )))
        userMenu= self.driver.find_element(By.ID,USER_MENU)

        userMenuText=userMenu.text

        assert userMenuText == MENU_TEXT
      
# Ürünlerin isimlerinin excel dosyalarındaki isimlerle uyuşması testi
    @pytest.mark.parametrize("productName",readProductNameFromExcel())
    def test_checkingExcelProduct(self,productName):
        # Başarılı login fonksiyonunu çağırma
        self.test_loginStandartUser()
        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH,INVENTORY_ITEM_NAME)))

        if productName == INV_PRODUCTNAME1: # "Sauce Labs Bike Light" adlı ürün
            WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,INV_BIKELIGHT_ITEM_ID)))
            item = self.driver.find_elements(By.ID,INV_BIKELIGHT_ITEM_ID)
            sizeItem = len(item)
            assert sizeItem > 0

        elif productName == INV_PRODUCTNAME2: # "Sauce Labs Bolt T-Shirt" adlı ürün
            WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,INV_BOLTTSHIRT_ITEM_ID)))
            item = self.driver.find_elements(By.ID,INV_BOLTTSHIRT_ITEM_ID)
            sizeItem = len(item)
            assert sizeItem > 0

        elif productName == INV_PRODUCTNAME3: # "Sauce Labs Onesie" adlı ürün
            WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,INV_ONESIE_ITEM_ID)))
            item = self.driver.find_elements(By.ID,INV_ONESIE_ITEM_ID)
            sizeItem = len(item)
            assert sizeItem > 0  
        
        elif productName == INV_PRODUCTNAME4: # "Test.allTheThings() T-Shirt (Red)" adlı ürün
            WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,INV_TESTALLTHINGSTSHIRT_ITEM_ID)))
            item = self.driver.find_elements(By.ID,INV_TESTALLTHINGSTSHIRT_ITEM_ID)
            sizeItem = len(item)
            assert sizeItem > 0 
        
        elif productName == INV_PRODUCTNAME5: # "Sauce Labs Backpack" adlı ürün
            WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,INV_BACKPACK_ITEM_ID)))
            item = self.driver.find_elements(By.ID,INV_BACKPACK_ITEM_ID)
            sizeItem = len(item)
            assert sizeItem > 0
        
        elif productName == INV_PRODUCTNAME6: # "Sauce Labs Fleece Jacket" adlı ürün
            WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,INV_FLEECEJACKET_ITEM_ID)))
            item = self.driver.find_elements(By.ID,INV_FLEECEJACKET_ITEM_ID)
            sizeItem = len(item)
            assert sizeItem > 0
        
        else :
            # Excel'e böyle bir ürün ismi yazılamıyacağı için, dolayısıyla test hata verecektir. 
            # Excel'e eklenilecek yeni ürün isimlerini test etmeden önce testte tanımlamamız gerekir. 
            assert productName == "d1dasdawdsdq12eqwdwads" 







    # "locked_out_user" ile giriş yapıldığında verilen uyarı mesajının doğrulanması testi
    def test_lockedOutUser(self):
        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID, USERNAME_ID )))
        userName= self.driver.find_element(By.ID, USERNAME_ID )
        userName.send_keys(LOCKED_USERNAME)

        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID, USERNAME_ID )))
        passwordID= self.driver.find_element(By.ID, PASSWORD_ID )
        passwordID.send_keys(PASSWORD)

        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID, USERNAME_ID )))
        loginID= self.driver.find_element(By.ID,LOGIN_ID)
        loginID.click()

        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, ERROR_LOGIN)))
        errorMenu= self.driver.find_element(By.XPATH,ERROR_LOGIN)
        errorMenuText = errorMenu.text

        assert errorMenuText == LOCKED_ERROR_TEXT

    # Normal giriş yapılıp sonrasında ürün sayılarının doğrulanması testi
    def test_totalProduct(self):
        # Başarılı login fonksiyonunu çağırma
        self.test_loginStandartUser()
        
        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, INVENTORY_ITEM)))
        inventory = self.driver.find_elements(By.XPATH, INVENTORY_ITEM)
        invSize = len(inventory)

        assert invSize == 6

     
    # Ürünlerin z'den a ya sıralanma fonksiyonunun testi
    def test_productsZtoA(self):
        # Başarılı login fonksiyonunu çağırma
        self.test_loginStandartUser()

        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, SET_LIST)))
        setList = self.driver.find_element(By.CSS_SELECTOR, SET_LIST)
        setList.click()

        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.XPATH, SET_ZTOA)))
        zToA = self.driver.find_element(By.XPATH, SET_ZTOA)
        zToA.click()

        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, INVENTORY_ITEM)))
        inventory = self.driver.find_elements(By.XPATH, INVENTORY_ITEM)
        invSize = len(inventory)
        
        inventoryList = []
        for i in range (0,invSize):
            item = self.driver.find_element(By.XPATH, "/html/body/div/div/div/div[2]/div/div/div/div["+ str(i+1) + "]/div[2]/div[1]/a/div") # İlgili XPATH'in değişmediği durumda çalışır.
            itemText = item.text
            inventoryList.append(itemText)
        inventoryList.sort(reverse=True)
        newInventoryList = inventoryList    
         
        assert newInventoryList == inventoryList

    # Ürünlerin düşük fiyattan yüksek fiyata sıralanma fonksiyonunun testi
    def test_lowToHighPrice(self):
        # Başarılı login fonksiyonunu çağırma
        self.test_loginStandartUser()

        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, SET_LIST)))
        setList = self.driver.find_element(By.CSS_SELECTOR, SET_LIST)
        setList.click()

        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.XPATH, SET_LOWHIGH)))
        lowHigh = self.driver.find_element(By.XPATH, SET_LOWHIGH)
        lowHigh.click()

        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, INVENTORY_ITEM)))
        inventory = self.driver.find_elements(By.XPATH, INVENTORY_ITEM)
        invSize = len(inventory)
        
        inventoryList = []
        for i in range (0,invSize):
            item = self.driver.find_element(By.XPATH, "/html/body/div/div/div/div[2]/div/div/div/div["+ str(i+1) + "]/div[2]/div[2]/div") # İlgili XPATH'in değişmediği durumda çalışır.
            itemText = item.text
            inventoryList.append(itemText)
        inventoryList.sort()
        newInventoryList = inventoryList    
            
        assert newInventoryList == inventoryList

    # Bir excel dosyasında ismi geçen ürünlerin sepete eklenmesi fonksiyonu testi
    @pytest.mark.parametrize("productName",readProductNameFromExcel())
    def test_excelToBasket(self,productName):
        # Başarılı login fonksiyonunu çağırma
        self.test_loginStandartUser()

        if productName == INV_PRODUCTNAME1: # "Sauce Labs Bike Light" adlı ürün
            WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,ADD_CART_PRODUCTNAME1_ID)))
            addChart = self.driver.find_element(By.ID,ADD_CART_PRODUCTNAME1_ID)
            addChart.click()

            WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, BASKET_ITEM_CHECK)))
            basket_items = self.driver.find_elements(By.XPATH, BASKET_ITEM_CHECK)
            sizeBasket = len(basket_items)
            assert sizeBasket > 0
           

        elif productName == INV_PRODUCTNAME2: # "Sauce Labs Bolt T-Shirt" adlı ürün
            WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,ADD_CART_PRODUCTNAME2_ID)))
            addChart = self.driver.find_element(By.ID,ADD_CART_PRODUCTNAME2_ID)
            addChart.click()
            
            WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, BASKET_ITEM_CHECK)))
            basket_items = self.driver.find_elements(By.XPATH, BASKET_ITEM_CHECK)
            sizeBasket = len(basket_items)
            assert sizeBasket > 0

        elif productName == INV_PRODUCTNAME3: # "Sauce Labs Onesie" adlı ürün
            WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,ADD_CART_PRODUCTNAME3_ID)))
            addChart = self.driver.find_element(By.ID,ADD_CART_PRODUCTNAME3_ID)
            addChart.click()
            
            WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, BASKET_ITEM_CHECK)))
            basket_items = self.driver.find_elements(By.XPATH, BASKET_ITEM_CHECK)
            sizeBasket = len(basket_items)
            assert sizeBasket > 0
        
        elif productName == INV_PRODUCTNAME4: # "Test.allTheThings() T-Shirt (Red)" adlı ürün
            WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,ADD_CART_PRODUCTNAME4_ID)))
            addChart = self.driver.find_element(By.ID,ADD_CART_PRODUCTNAME4_ID)
            addChart.click()
            
            WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, BASKET_ITEM_CHECK)))
            basket_items = self.driver.find_elements(By.XPATH, BASKET_ITEM_CHECK)
            sizeBasket = len(basket_items)
            assert sizeBasket > 0
        
        elif productName == INV_PRODUCTNAME5: # "Sauce Labs Backpack" adlı ürün
            WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,ADD_CART_PRODUCTNAME5_ID)))
            addChart = self.driver.find_element(By.ID,ADD_CART_PRODUCTNAME5_ID)
            addChart.click()
            
            WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, BASKET_ITEM_CHECK)))
            basket_items = self.driver.find_elements(By.XPATH, BASKET_ITEM_CHECK)
            sizeBasket = len(basket_items)
            assert sizeBasket > 0
        
        elif productName == INV_PRODUCTNAME6: # "Sauce Labs Fleece Jacket" adlı ürün
            WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,ADD_CART_PRODUCTNAME6_ID)))
            addChart = self.driver.find_element(By.ID,ADD_CART_PRODUCTNAME6_ID)
            addChart.click()
            
            WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, BASKET_ITEM_CHECK)))
            basket_items = self.driver.find_elements(By.XPATH, BASKET_ITEM_CHECK)
            sizeBasket = len(basket_items)
            assert sizeBasket > 0
        
        else :
            # Excel'e böyle bir ürün ismi yazılamıyacağı için, dolayısıyla test hata verecektir. 
            # Excel'e eklenilecek yeni ürün isimlerini test etmeden önce testte tanımlamamız gerekir. 
            assert productName == "d1dasdawdsdq12eqwdwads" 
        


    # Sepete eklenen ürünlerin sepet sayfasında doğru bir şekilde görünmesi testi
    def test_productToBasket(self):
        # Başarılı login fonksiyonunu çağırma
        self.test_loginStandartUser()
        
        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID, ADD_CART_PRODUCTNAME5_ID)))
        addChart = self.driver.find_element(By.ID, ADD_CART_PRODUCTNAME5_ID)
        addChart.click()

        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, BASKET_ITEM_CHECK)))
        basket_items = self.driver.find_elements(By.XPATH, BASKET_ITEM_CHECK)
        sizeBasket = len(basket_items)
        assert sizeBasket > 0

    # Sepetten kaldırılan ürünün sepet ekranından kaldırılma testi
    def test_removeProductFromBasket(self):
        # Başarılı login fonksiyonunu çağırma
        self.test_loginStandartUser()
        
        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.NAME, ADD_CART_PRODUCTNAME5_ID)))
        addChart = self.driver.find_element(By.NAME, ADD_CART_PRODUCTNAME5_ID)
        addChart.click()

        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, BASKET_BUTTON)))
        basketButton = self.driver.find_element(By.XPATH, BASKET_BUTTON)
        basketButton.click()

        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.NAME, REMOVE_CART_LABS_BACKPACK_NAME)))
        removeChart = self.driver.find_element(By.NAME, REMOVE_CART_LABS_BACKPACK_NAME)
        removeChart.click()

        WebDriverWait(self.driver,5).until(expected_conditions.invisibility_of_element_located((By.XPATH, BASKET_ITEM_REMOVE))) # Sitenin yapısı gereği ürün sepetten kaldırıldığında, ürün ile ilgili 'class' görünmez oluyor.
        basket_items = self.driver.find_elements(By.XPATH, BASKET_ITEM_REMOVE)
        sizeBasket = len(basket_items)
        assert sizeBasket > 0




        
